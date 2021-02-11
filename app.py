from openpyxl import Workbook, load_workbook
import pandas as pd

while(True):
    print("Welcome to Munder Difflin Paper")
    wb = load_workbook('Munder Difflin.xlsx')
    cust_no = 1
    sheet = wb.active

    client_email = input('Enter the client\'s email address: ')
    client_name = input('Enter the client\'s name: ')
    client_address = input('Enter the client\'s address: ')
    description_sale = input('Enter the sale description: ')
    cost = input('Enter the cost of sale: ')
    inv_number = input('Enter the invoice number: ')
    today_date = input('Enter today\'s date: ')
    due_date = input('Enter the due date for  the invoice: ')


    sheet.cell(row=10, column=2).value = client_name
    sheet.cell(row=10, column=4).value = client_address
    sheet.cell(row=9, column=6).value = inv_number
    sheet.cell(row=10, column=6).value = today_date
    sheet.cell(row=11, column=6).value = due_date
    sheet.cell(row=16, column=2).value = description_sale
    sheet.cell(row=16, column=5).value = cost

    filename = 'Dunds'+str(cust_no)+'.xlsx'
    wb.save(filename)
    cust_no = cust_no +1 
    print("Do you want to create another one?")
    choice = input('Do you want to create another one? ')
    if (int(choice)) == 1:
        pass
    else:
        break
